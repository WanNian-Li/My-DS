import argparse
import json
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def normalize_scene_name(name: str) -> str:
    """Convert scene file name to scene id used in Excel cells."""
    text = str(name).strip()
    if text.lower().endswith(".nc"):
        return text[:-3]
    return text


def load_scene_names(json_path: Path) -> list[str]:
    with json_path.open("r", encoding="utf-8") as f:
        data = json.load(f)

    if not isinstance(data, list):
        raise ValueError(f"JSON must be a list, got {type(data).__name__}")

    names = []
    for item in data:
        if not isinstance(item, str):
            continue
        normalized = normalize_scene_name(item)
        if normalized:
            names.append(normalized)

    return names


def highlight_matches(excel_path: Path, targets: set[str], output_path: Path) -> tuple[dict[str, list[str]], set[str]]:
    wb = load_workbook(excel_path)
    fill = PatternFill(fill_type="solid", fgColor="4C2872")  # light yellow

    matched_cells: dict[str, list[str]] = {}

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str):
                    continue

                value = cell.value.strip()
                if value in targets:
                    cell.fill = fill
                    matched_cells.setdefault(value, []).append(f"{ws.title}!{cell.coordinate}")

    wb.save(output_path)

    matched_names = set(matched_cells.keys())
    missing_names = targets - matched_names
    return matched_cells, missing_names


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Match scene names from JSON in Excel and highlight matched cells."
    )
    parser.add_argument(
        "--json",
        default="datalists/val_list.json",
        help="Path to JSON list of scene file names (.nc allowed).",
    )
    parser.add_argument(
        "--excel",
        default=r"F:\\ZJU\\11_Ice\\dataset_create\\train_scene_metrics.xlsx",
        help="Path to source Excel file.",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Path to output Excel file. Default: <excel_name>_highlighted.xlsx",
    )
    return parser


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()

    json_path = Path(args.json)
    excel_path = Path(args.excel)

    if not json_path.exists():
        raise FileNotFoundError(f"JSON file not found: {json_path}")
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    output_path = Path(args.output) if args.output else excel_path.with_name(f"{excel_path.stem}_highlighted{excel_path.suffix}")

    names = load_scene_names(json_path)
    targets = set(names)

    if not targets:
        raise ValueError("No valid scene names found in JSON.")

    matched_cells, missing_names = highlight_matches(excel_path, targets, output_path)

    total_matches = sum(len(v) for v in matched_cells.values())
    print(f"Input names: {len(targets)}")
    print(f"Matched names: {len(matched_cells)}")
    print(f"Matched cells: {total_matches}")
    print(f"Output file: {output_path}")

    if missing_names:
        print("\nNames not found in Excel:")
        for name in sorted(missing_names):
            print(name)


if __name__ == "__main__":
    main()
